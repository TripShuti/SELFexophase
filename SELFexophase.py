import json
import csv
import requests
from pathlib import Path
from datetime import datetime, timezone
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from textual import work
from textual.app import App, ComposeResult
from textual.screen import ModalScreen
from textual.containers import Vertical, Horizontal
from textual.widgets import (
    Header, Footer, DataTable, TabbedContent, TabPane,
    Input, Button, Label, Static
)

OUTPUT_DIR = "exophase_json"

PLATFORM_RULES = [
    ("PlayStation", ["playstation", "ps4", "ps5", "ps3", "ps vita"]),
    ("Xbox",        ["xbox"]),
    ("Steam",       ["steam"]),
]

def classify_platform(platforms: list[dict]) -> str:
    names_lower = [str(p.get("name", "")).lower() for p in platforms if isinstance(p, dict)]
    for tab_name, keywords in PLATFORM_RULES:
        if any(kw in name for name in names_lower for kw in keywords):
            return tab_name
    return "Other"

def fmt_playtime(game: dict) -> str:
    units = game.get("playtimeUnits") or {}
    h, m = units.get("hours", 0), units.get("minutes", 0)
    return f"{h}h {m}m" if (h or m) else str(game.get("playtime", "0h"))

def fmt_timestamp(ts) -> str:
    if ts and ts > 0:
        try:
            return datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    return ""

def build_row(game: dict, platform: str) -> tuple:
    meta       = game.get("meta") or {}
    title      = meta.get("title", "Unknown")
    playtime   = fmt_playtime(game)
    percent    = game.get("percent", 0.0)
    lastplayed = fmt_timestamp(game.get("lastplayed_utc", 0))

    if platform == "PlayStation":
        return (
            title, playtime,
            game.get("earned_bronze", 0) or 0,
            game.get("earned_silver", 0) or 0,
            game.get("earned_gold", 0) or 0,
            game.get("earned_platinum", 0) or 0,
            percent, lastplayed,
        )
    if platform == "Xbox":
        return (
            title, playtime,
            game.get("earned_awards", 0) or 0,
            game.get("total_awards", 0) or 0,
            game.get("earned_points", 0) or 0,
            percent, lastplayed,
        )
    if platform == "Steam":
        return (
            title, playtime,
            game.get("earned_awards", 0) or 0,
            game.get("total_awards", 0) or 0,
            percent, lastplayed,
        )
    platforms_str = ", ".join(
        str(p.get("name", "")) for p in (meta.get("platforms") or []) if isinstance(p, dict)
    )
    return (title, playtime, platforms_str, percent, lastplayed)


MODAL_CSS = """
ModalScreen {
    align: center middle;
    background: $background 80%;
}
#dialog {
    width: 64;
    height: auto;
    padding: 2 3;
    background: $surface;
    border: solid $accent;
}
#dialog Label { margin-bottom: 1; text-style: bold; }
#dialog Input  { margin-bottom: 1; }
Horizontal { margin-top: 1; height: auto; align: right middle; }
Button { margin-left: 1; }
"""

class SyncModal(ModalScreen[str | None]):
    CSS = MODAL_CSS

    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label("Синхронізація з Exophase")
            yield Input(placeholder="Player ID", id="pid")
            with Horizontal():
                yield Button("Скасувати", variant="error",   id="cancel")
                yield Button("Синхронізувати", variant="success", id="start")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "start":
            pid = self.query_one("#pid", Input).value.strip()
            if pid:
                self.dismiss(pid)
            else:
                self.app.notify("Введіть Player ID!", severity="warning")
        else:
            self.dismiss(None)


class FilterModal(ModalScreen[str | None]):
    CSS = MODAL_CSS

    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label("Фільтр за назвою гри")
            yield Input(placeholder="Введіть частину назви…", id="query")
            with Horizontal():
                yield Button("Скинути", variant="warning", id="reset")
                yield Button("Скасувати", variant="error", id="cancel")
                yield Button("Застосувати", variant="success", id="apply")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "apply":
            self.dismiss(self.query_one("#query", Input).value.strip())
        elif event.button.id == "reset":
            self.dismiss("")
        else:
            self.dismiss(None)


TABLE_COLUMNS = {
    "dt-ps":    ("Game", "Playtime", "Bronze", "Silver", "Gold", "Platinum", "Completion %", "Last Played (UTC)"),
    "dt-xbox":  ("Game", "Playtime", "Earned Awards", "Total Awards", "Earned Points", "Completion %", "Last Played (UTC)"),
    "dt-steam": ("Game", "Playtime", "Earned Awards", "Total Awards", "Completion %", "Last Played (UTC)"),
    "dt-other": ("Game", "Playtime", "Platforms", "Completion %", "Last Played (UTC)"),
}

TAB_TO_DT = {
    "tp-ps":    "#dt-ps",
    "tp-xbox":  "#dt-xbox",
    "tp-steam": "#dt-steam",
    "tp-other": "#dt-other",
}

PLATFORM_TO_DT = {
    "PlayStation": "#dt-ps",
    "Xbox":        "#dt-xbox",
    "Steam":       "#dt-steam",
    "Other":       "#dt-other",
}


class SELFexophase(App):
    CSS = """
    DataTable { width: 1fr; height: 1fr; }
    #statusbar { height: 1; background: $surface; color: $text-muted; padding: 0 2; }
    """
    BINDINGS = [
        ("q", "quit",        "Вихід"),
        ("s", "sync",        "Синхронізувати"),
        ("d", "delete_game", "Видалити"),
        ("e", "export",      "Експорт Excel"),
        ("c", "export_csv",  "Експорт CSV"),
        ("f", "filter",      "Фільтр"),
        ("r", "reload",      "Перезавантажити"),
    ]

    def __init__(self):
        super().__init__()
        self.sort_state:      dict[str, bool] = {}
        self.games_data:      list[dict]      = []
        self.latest_json_path: Path | None    = None
        self._filter_query:   str             = ""

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with TabbedContent():
            with TabPane("PlayStation", id="tp-ps"):
                yield DataTable(id="dt-ps")
            with TabPane("Xbox", id="tp-xbox"):
                yield DataTable(id="dt-xbox")
            with TabPane("Steam", id="tp-steam"):
                yield DataTable(id="dt-steam")
            with TabPane("Other", id="tp-other"):
                yield DataTable(id="dt-other")
        yield Static("", id="statusbar")
        yield Footer()

    def on_mount(self) -> None:
        for dt_id, cols in TABLE_COLUMNS.items():
            dt = self.query_one(f"#{dt_id}", DataTable)
            dt.cursor_type  = "row"
            dt.zebra_stripes = True
            dt.add_columns(*cols)
        self.load_data()

    def load_data(self, filter_query: str = "") -> None:
        for dt_id in TABLE_COLUMNS:
            self.query_one(f"#{dt_id}", DataTable).clear()

        base_dir = Path(__file__).resolve().parent
        files = list(base_dir.rglob("all_games_*.json"))
        if not files:
            self.games_data = []
            self._update_statusbar()
            return

        self.latest_json_path = max(files, key=lambda p: p.stat().st_mtime)
        try:
            with open(self.latest_json_path, "r", encoding="utf-8") as f:
                self.games_data = json.load(f)
        except Exception as e:
            self.notify(f"Помилка читання JSON: {e}", severity="error")
            return

        tables = {k: self.query_one(v, DataTable) for k, v in PLATFORM_TO_DT.items()}
        q = filter_query.lower()
        loaded = 0

        for game in self.games_data:
            if not isinstance(game, dict):
                continue
            meta  = game.get("meta") or {}
            title = meta.get("title", "Unknown")
            if q and q not in title.lower():
                continue

            platforms = meta.get("platforms") or []
            platform  = classify_platform(platforms)
            row       = build_row(game, platform)
            tables[platform].add_row(*row)
            loaded += 1

        indicator = f"  🔍 фільтр: «{filter_query}»" if filter_query else ""
        self._update_statusbar(loaded, indicator)

    def _update_statusbar(self, count: int = 0, extra: str = "") -> None:
        file_info = self.latest_json_path.name if self.latest_json_path else "—"
        self.query_one("#statusbar", Static).update(
            f" 📁 {file_info}   |   🎮 {count} ігор{extra}"
        )

    @staticmethod
    def _sort_key(value) -> tuple:
        if isinstance(value, (int, float)):
            return (0, float(value))
        if isinstance(value, str):
            if not value:
                return (0, 0.0)
            if "h " in value and "m" in value:
                try:
                    h, rest = value.split("h ")
                    m = rest.replace("m", "").strip()
                    return (0, int(h) * 60 + int(m))
                except ValueError:
                    pass
            try:
                return (0, float(value))
            except ValueError:
                return (1, value.lower())
        return (1, str(value).lower())

    def on_data_table_header_selected(self, event: DataTable.HeaderSelected) -> None:
        dt        = event.data_table
        col_key   = event.column_key
        state_key = f"{dt.id}_{col_key}"
        reverse   = self.sort_state.get(state_key, False)
        dt.sort(col_key, key=self._sort_key, reverse=reverse)
        self.sort_state[state_key] = not reverse

    def _active_datatable(self) -> DataTable | None:
        active = self.query_one(TabbedContent).active
        dt_selector = TAB_TO_DT.get(active)
        return self.query_one(dt_selector, DataTable) if dt_selector else None

    def _output_dir(self) -> Path:
        d = Path(__file__).resolve().parent / OUTPUT_DIR
        d.mkdir(exist_ok=True)
        return d

    def _save_games_data(self) -> None:
        if not self.latest_json_path:
            self.latest_json_path = self._output_dir() / "all_games_manual.json"
        try:
            with open(self.latest_json_path, "w", encoding="utf-8") as f:
                json.dump(self.games_data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            self.notify(f"Помилка збереження: {e}", severity="error")

    def action_reload(self) -> None:
        self.load_data(self._filter_query)
        self.notify("Дані перезавантажено", severity="information")

    def action_filter(self) -> None:
        def apply_filter(query: str | None) -> None:
            if query is None:
                return
            self._filter_query = query
            self.load_data(query)
            if query:
                self.notify(f"Фільтр: «{query}»", severity="information")
            else:
                self.notify("Фільтр скинуто", severity="information")

        self.push_screen(FilterModal(), apply_filter)

    def action_delete_game(self) -> None:
        dt = self._active_datatable()
        if dt is None:
            return
        try:
            row_key    = dt.coordinate_to_cell_key(dt.cursor_coordinate).row_key
            row_values = dt.get_row(row_key)
            target_title = str(row_values[0])
        except Exception:
            self.notify("Немає виділеного рядка", severity="warning")
            return

        for i, game in enumerate(self.games_data):
            meta = game.get("meta") or {}
            if meta.get("title", "") == target_title:
                del self.games_data[i]
                break

        self._save_games_data()
        self.load_data(self._filter_query)
        self.notify(f"Видалено: {target_title}", severity="information")

    def action_sync(self) -> None:
        def on_result(player_id: str | None) -> None:
            if player_id:
                self.notify(f"Синхронізація гравця «{player_id}»…", severity="information")
                self.fetch_api_data(player_id)

        self.push_screen(SyncModal(), on_result)

    def action_export(self) -> None:
        self.notify("Формування Excel файлу…", severity="information")
        self.export_excel_data()

    def action_export_csv(self) -> None:
        self.notify("Формування CSV файлу…", severity="information")
        self.export_csv_data()

    @work(thread=True)
    def fetch_api_data(self, player_id: str) -> None:
        out_dir = self._output_dir()
        headers = {
            "User-Agent":      "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                               "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept":          "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer":         "https://www.exophase.com/",
        }

        all_games: list[dict] = []
        for page_num in range(1, 9999):
            url = (
                f"https://api.exophase.com/public/player/{player_id}/games"
                f"?page={page_num}&environment=&sort=1&showHidden=0"
            )
            try:
                resp = requests.get(url, headers=headers, timeout=10)
                resp.raise_for_status()
                data = resp.json()
            except Exception as e:
                self.call_from_thread(self.notify, f"Помилка API (стор. {page_num}): {e}", severity="error")
                break

            if not data.get("success"):
                break
            games = data.get("games") or []
            if not games:
                break
            all_games.extend(games)

        if not all_games:
            self.call_from_thread(self.notify, "Нічого не отримано від API", severity="warning")
            return

        json_path = out_dir / f"all_games_{player_id}.json"
        try:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(all_games, f, ensure_ascii=False, indent=4)
            self.call_from_thread(
                self.notify, f"✅ Синхронізовано {len(all_games)} ігор → {json_path.name}", severity="information"
            )
            self.call_from_thread(self.load_data, self._filter_query)
        except Exception as e:
            self.call_from_thread(self.notify, f"Помилка збереження: {e}", severity="error")

    @work(thread=True)
    def export_excel_data(self) -> None:
        out_dir = self._output_dir()
        if not self.games_data:
            self.call_from_thread(self.notify, "Немає даних для експорту", severity="error")
            return

        sheets: dict[str, list[dict]] = {k: [] for k in ("PlayStation", "Xbox", "Steam", "Other")}

        for game in self.games_data:
            meta      = game.get("meta") or {}
            platforms = meta.get("platforms") or []
            platform  = classify_platform(platforms)
            row_tuple = build_row(game, platform)
            col_names = TABLE_COLUMNS[PLATFORM_TO_DT[platform].lstrip("#")]
            sheets[platform].append(dict(zip(col_names, row_tuple)))

        player_id   = (self.latest_json_path.stem.split("_")[-1]
                       if self.latest_json_path else "export")
        excel_path  = out_dir / f"exophase_games_{player_id}.xlsx"

        HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
        HEADER_FONT  = Font(color="FFFFFF", bold=True)
        HEADER_ALIGN = Alignment(horizontal="center")

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                for sheet_name, rows in sheets.items():
                    if not rows:
                        continue
                    df = pd.DataFrame(rows)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]

                    for cell in ws[1]:
                        cell.fill      = HEADER_FILL
                        cell.font      = HEADER_FONT
                        cell.alignment = HEADER_ALIGN

                    for col_idx, col in enumerate(ws.columns, 1):
                        width = max((len(str(c.value)) for c in col if c.value), default=8)
                        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

            self.call_from_thread(
                self.notify, f"✅ Excel → {excel_path.name}", severity="information"
            )
        except Exception as e:
            self.call_from_thread(self.notify, f"Помилка Excel-експорту: {e}", severity="error")

    @work(thread=True)
    def export_csv_data(self) -> None:
        if not self.games_data:
            self.call_from_thread(self.notify, "Немає даних для CSV-експорту", severity="error")
            return

        out_dir   = self._output_dir()
        tabs      = self.query_one(TabbedContent)
        active_id = tabs.active
        platform  = active_id.replace("tp-", "").capitalize()
        if platform == "Ps":
            platform = "PlayStation"

        rows: list[dict] = []
        for game in self.games_data:
            meta  = game.get("meta") or {}
            plat  = classify_platform(meta.get("platforms") or [])
            if plat.lower() != platform.lower():
                continue
            row_tuple = build_row(game, plat)
            col_names = TABLE_COLUMNS[PLATFORM_TO_DT[plat].lstrip("#")]
            rows.append(dict(zip(col_names, row_tuple)))

        if not rows:
            self.call_from_thread(self.notify, f"Немає даних для вкладки {platform}", severity="warning")
            return

        player_id = (self.latest_json_path.stem.split("_")[-1]
                     if self.latest_json_path else "export")
        csv_path  = out_dir / f"exophase_{platform.lower()}_{player_id}.csv"

        try:
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
            self.call_from_thread(
                self.notify, f"✅ CSV → {csv_path.name}", severity="information"
            )
        except Exception as e:
            self.call_from_thread(self.notify, f"Помилка CSV-експорту: {e}", severity="error")


if __name__ == "__main__":
    SELFexophase().run()